from odoo import models, fields, _, api
from odoo.exceptions import ValidationError
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import base64
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
class EmployeeSalaryData(models.Model):
    _name = 'employee.salary.data'
    _description = 'Employee Salary Data'

    employee_id = fields.Many2one(
        'hr.employee',
        string='Employee',
        required=True,
    )

    date = fields.Date(
        string='Salary Date',
        required=True,
        default=fields.Date.today,
    )

    basic_salary = fields.Float(
        string='Basic Salary',
        digits=(10, 2)
    )

    working_days = fields.Float(
        string='Working Days',
        digits=(10, 2)
    )

    overtime = fields.Float(
        string='Overtime',
        digits=(10, 2)
    )

    net_salary = fields.Float(
        string='Net Salary',
        compute='_compute_net_salary',
        digits=(10, 2)
    )

    @api.depends('basic_salary', 'overtime')
    def _compute_net_salary(self):
        for record in self:
            record.net_salary = record.basic_salary + record.overtime


class EmployeeSalaryWizard(models.TransientModel):
    _name = 'employee.salary.wizard'
    _description = 'Employee Salary Selection Wizard'

    employee_ids = fields.Many2many('hr.employee', string='Employees')
    selected_fields = fields.Many2many(
        'ir.model.fields',
        domain="[('model', '=', 'employee.salary.data'), "
               "('name', 'in', ['basic_salary', 'working_days', 'overtime', 'net_salary'])]",
        string='Select Fields',
        required=True
    )

    def action_confirm(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Salary Report'

        # Set page layout
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToPage = True

        # Add company logo with better positioning
        img = Image(
            '/home/tarek/PycharmProjects/odoo17/odoo17/odoo/custom_addons/dynamic_employee_excel_report/models/img.jpg')
        img.width = 75
        img.height = 38
        sheet.add_image(img, 'H1')

        # Create a larger merged section for the title
        sheet.merge_cells('B1:H2')
        title_cell = sheet['B1']
        title_cell.value = 'Employee Salary Report'
        title_cell.font = Font(name='Calibri', size=25, bold=True, color="1F497D")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add company info section
        sheet.merge_cells('A3:B3')
        sheet['A3'] = 'Report Generated:'
        sheet['A3'].font = Font(bold=True)
        sheet['C3'] = fields.Date.today().strftime('%d-%m-%Y')

        # Add decorative line under header
        for col in range(1, 9):  # Adjust range based on your columns
            cell = sheet.cell(row=4, column=col)
            cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
            cell.border = Border(bottom=Side(style='thick', color='1F497D'))

        # Start data from row 6 to give more space for header
        data_start_row = 6

        # Freeze panes for better navigation
        sheet.freeze_panes = f'A{data_start_row + 1}'

        # Set professional column widths
        sheet.column_dimensions['A'].width = 35
        for col, field in enumerate(self.selected_fields, start=1):
            column_letter = get_column_letter(col + 1)
            sheet.column_dimensions[column_letter].width = 18

        # Enhanced header style with gradient-like effect
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Write headers with enhanced style
        headers = ['Employee Name'] + [field.field_description for field in self.selected_fields]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=data_start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='B4C6E7'),
                right=Side(style='thin', color='B4C6E7'),
                top=Side(style='thin', color='B4C6E7'),
                bottom=Side(style='thin', color='B4C6E7')
            )

        # Professional alternating row colors
        row_colors = ['EDF3F9', 'FFFFFF']  # Light blue and white

        # Write data with enhanced formatting
        row = data_start_row + 1
        employees = self.employee_ids if self.employee_ids else self.env['hr.employee'].search([])

        for index, employee in enumerate(employees):
            # Alternate row colors
            row_color = row_colors[index % 2]
            row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

            salary_record = self.env['employee.salary.data'].search([
                ('employee_id', '=', employee.id)
            ], limit=1)

            # Employee name cell
            name_cell = sheet.cell(row=row, column=1, value=employee.name)
            name_cell.fill = row_fill
            name_cell.alignment = Alignment(horizontal='left')
            name_cell.border = Border(
                left=Side(style='thin', color='B4C6E7'),
                right=Side(style='thin', color='B4C6E7'),
                top=Side(style='thin', color='B4C6E7'),
                bottom=Side(style='thin', color='B4C6E7')
            )

            # Data cells
            for col, field in enumerate(self.selected_fields, start=2):
                cell = sheet.cell(row=row, column=col)
                value = salary_record[field.name] if salary_record else 0

                cell.fill = row_fill
                cell.border = Border(
                    left=Side(style='thin', color='B4C6E7'),
                    right=Side(style='thin', color='B4C6E7'),
                    top=Side(style='thin', color='B4C6E7'),
                    bottom=Side(style='thin', color='B4C6E7')
                )
                cell.alignment = Alignment(horizontal='center')

                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                    if value < 0:
                        cell.font = Font(color="FF0000")
                else:
                    cell.value = value

            row += 1

        # Enhanced Summary Section
        summary_row = row + 3  # Add more space before summary

        # Add decorative line above summary
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=summary_row - 1, column=col)
            cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')

        # Summary header
        sheet.merge_cells(f'A{summary_row}:B{summary_row}')
        summary_header = sheet[f'A{summary_row}']
        summary_header.value = 'Statistical Summary'
        summary_header.font = Font(name='Calibri', size=14, bold=True, color='1F497D')
        summary_header.alignment = Alignment(horizontal='left')

        # Statistics with better formatting
        stats_row = summary_row + 1
        stats_labels = ['Total', 'Average', 'Maximum']

        for idx, label in enumerate(stats_labels):
            row_num = stats_row + idx
            label_cell = sheet[f'A{row_num}']
            label_cell.value = label
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='right')

            for col, field in enumerate(self.selected_fields, start=1):
                column_letter = get_column_letter(col + 1)
                cell = sheet.cell(row=row_num, column=col + 1)

                if label == 'Total':
                    cell.value = f'=SUM({column_letter}{data_start_row + 1}:{column_letter}{row - 1})'
                elif label == 'Average':
                    cell.value = f'=AVERAGE({column_letter}{data_start_row + 1}:{column_letter}{row - 1})'
                else:  # Maximum
                    cell.value = f'=MAX({column_letter}{data_start_row + 1}:{column_letter}{row - 1})'

                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True if label == 'Total' else False)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin', color='B4C6E7'),
                    right=Side(style='thin', color='B4C6E7'),
                    top=Side(style='thin', color='B4C6E7'),
                    bottom=Side(style='thin', color='B4C6E7')
                )

        # Add chart with better positioning and style
        chart_row = stats_row + 5
        chart = BarChart()
        chart.title = "Salary Distribution Overview"
        chart.style = 42  # Professional style
        chart.height = 12
        chart.width = 28
        chart.y_axis.title = 'Amount'
        chart.x_axis.title = 'Employee'

        chart.graphicalProperties = GraphicalProperties(
            solidFill=ColorChoice(srgbClr='F0F8FF')  # Light gray background
        )

        # Set plot area background color
        chart.plot_area.graphicalProperties = GraphicalProperties(
            solidFill=ColorChoice(srgbClr='FFFFFF')  # White background for plot area
        )
        # Enhanced chart colors - Simplified version
        data = Reference(sheet, min_col=2, min_row=data_start_row,
                         max_col=len(headers), max_row=row - 1)
        cats = Reference(sheet, min_col=1, min_row=data_start_row + 1, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Set vibrant colors for each series
        colors = ['4472C4', 'ED7D31', '70AD47', 'FFC000', '5B9BD5', 'A5A5A5']
        for i, series in enumerate(chart.series):
            series.graphicalProperties.solidFill = colors[i % len(colors)]

        # Customize chart appearance
        chart.legend.position = 'r'  # Position legend on right
        chart.legend.horz = False  # Vertical legend

        # Add data labels
        for series in chart.series:
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True  # Show values
            series.dLbls.showCatName = False  # Don't show category names
            series.dLbls.showSerName = False  # Don't show series names

        # Add the chart with specific positioning
        sheet.add_chart(chart, f"A{chart_row}")

        # Save and return Excel file
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_binary = base64.b64encode(excel_file.getvalue())

        attachment = self.env['ir.attachment'].create({
            'name': f'enhanced_salary_report_{fields.Date.today()}.xlsx',
            'type': 'binary',
            'datas': excel_binary,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
